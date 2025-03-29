#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''=================================================
@Author ：kw
@IDE    ：PyCharm
@Date   ：2023/11/4 14:58
=================================================='''
import openpyxl

def read1(category_name):
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
    # sheets = file1.sheetnames
    # print(sheets)  # ['sheet', '妈妈用品', '产房用品', '住院用品', '需带证件']
    sheet = file1["妈妈用品"]
    # print(sheet)  # <Worksheet "妈妈用品">
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[(cell.column) - 1].value == category_name:
                m = i[4].value
                sum = sum + m
    print('read1函数的运行结果：' + str(sum))  # 533.96
# category = '护理类'
cell_id = int(input('请输入类别编号（0~哺乳类用品,1~护理类，2~衣帽类，3~其他生活用品）：'))
category = ''
if cell_id == 0:
    category = '哺乳类用品'
elif cell_id == 1:
    category = '护理类'
elif cell_id == 2:
    category = '衣帽类'
else:
    cell_id = '其他生活用品'
read1(category)

def read2():
    '''
    1、打开excel文件
    :return:
    '''
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')

    '''
    2、找到对应的sheet表
    '''
    sheet = file1["妈妈用品"]  # 获取表单
    # print(sheet.cell(row=1, column=1).value)  # 输出表单第一行第一列的值

    sum = 0
    for i in sheet.iter_rows():  # 遍历sheet表（妈妈用品）中的每一行
        for cell in i:  # 遍历每一行中的每一个单元格
            if i[(cell.column) - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read2函数的运行结果：' + str(sum))
read2()

def read3():
    try:
        file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
        sheet = file1["妈妈用品"]

        sum = 0
        for i in sheet.iter_rows():  # 遍历sheet表（妈妈用品）中的每一行
            for cell in i:  # 遍历每一行中的每一个单元格
                if i[(cell.column) - 1].value == '护理类':
                    m = i[4].value
                    sum = sum + m
        print('read3函数的运行结果：' + str(sum))
    except OSError as e:
        print('无效的文件' + e)
read3()

def read4():
    try:
        file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
        sheet = file1['妈妈用品']
        sum = 0

        for i in sheet.iter_rows():
            for cell in i:
                if i[(cell.column) - 1].value == '护理类':
                    m = i[4].value
                    sum = sum + m
        print('read4函数的运行结果：' + str(sum))
    except OSError as e:
        print('无效文件')
read4()

def read5():
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file1['妈妈用品']

    sum = 0
    for i in sheet.iter_rows():
        for c in i:
            if i[c.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read5函数的运行结果：' + str(sum))
read5()

def read6():
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file1['妈妈用品']

    sum = 0

    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read6函数的运行结果：' + str(sum))
read6()

def read7():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']

    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read7函数的运行结果：' + str(sum))
read7()

def read8():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read8函数的运行结果：' + str(sum))
read8()

def read9():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read9函数的运行结果：' + str(sum))
read9()

def read10():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read10函数的运行结果：' + str(sum))
read10()

def read11():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read11函数的运行结果：' + str(sum))
read11()

def read12():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read12函数的运行结果：' + str(sum))
read12()

def read13():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read13函数的运行结果：' + str(sum))
read13()

def read14():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m

    print('read14函数的运行结果：' + str(sum))
read14()

def read15():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read15函数的运行结果：' + str(sum))
read15()

def read16():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print('read16函数的运行结果：' + str(sum))
read16()

def read17():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']

    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)
read17()

def read18():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)
read18()

def read19():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)
read19()

def read20():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)
read20()

def read21():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)
read21()

def read22():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)
read22()

def read23():
    file=openpyxl.load_workbook('../resources/da')
    sheet=file['妈妈用品']

    sum=0
    for i in sheet.iter_rows():
        for cell in i:
            if i[cell.column-1].value=='护理类':
                m=i[4].value
                sum=sum+m
    print(sum)
read23()