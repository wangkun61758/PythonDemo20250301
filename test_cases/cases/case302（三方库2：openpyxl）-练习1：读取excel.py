#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''=================================================
@Author ：kw
@IDE    ：PyCharm
@Date   ：2023/10/18 20:43
=================================================='''
import openpyxl
'''
1、导入对应的库 —— openpyxl
'''
def test1():
    try:
        '''
        1、打开 Excel 文件
        2、注意：
        （1）openpyxl不支持旧版的xls格式，需要使用xlsx格式
        （2)读取的是xls文件不能是通过改变后缀变成xlsx格式的文件（比如直接将A.xls修改名称为A.xlsx）。
        不然读取文件会报错：openpyxl报错修改：OSError: File contains no valid workbook part
        '''
        file = openpyxl.load_workbook("../../resources/data1.xlsx")
        # 2、选择工作表
        worksheet = file['妈妈用品']
        # 3、指定单元格
        cell = worksheet.cell(row=1, column=2)

        # 4、读取单元格值
        cell_value = cell.value
        print("单元格值:", cell_value)

        # 5、关闭工作簿
        file.close()

    except OSError as e:
        print("无效的 Excel 文件:", str(e))
def test2():
    file = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file['妈妈用品']

    # 目标值
    target_value = '60片/盒'

    # 3.1、遍历工作表中的每一行
    for i in sheet.iter_rows():
        # 3.2、遍历每一行中的每一个单元格
        for cell in i:
            # 3.3、判断单元格的值是否等于目标值
            if cell.value == target_value:
                # 3.4、获取行和列的信息
                row_index = i[1].row
                column_index = cell.column
                # 输出结果
                print("目标值所在的行和列分别为：", row_index, column_index)  # 11 3
def test3():
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet1 = file1['妈妈用品']
    for i in sheet1.iter_rows():
        print(i[1].value)
    '''
品牌1
十月结晶 - 护理垫+刀纸
十月结晶 - 计量卫生巾
十月结晶 - 安睡裤
十月结晶 - 护理垫
十月结晶 - 一次性纯棉内裤
高洁丝 - 日用卫生巾
高洁丝 - 夜用卫生巾1
高洁丝 - 夜用卫生巾2
十月结晶 - 私处清洗器
洁丽雅 - 一次性马桶坐垫
维达 - 普通湿巾
维达 - 酒精湿巾
matyz - 硅胶吸奶器
新贝 - 防溢乳垫
兰思诺 - 羊毛脂膏
嫚熙 - 夏季纯棉两件套睡衣（带哺乳口）
嫚熙 - 四季款纯棉两件套睡衣（带哺乳口）
嫚熙 - 哺乳内衣
嫚熙 - 夏季纯棉月子帽
十月结晶 - 月子袜
克洛伊 - 吸管水杯
None
None
None
None

    '''
def test4():
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet1 = file1['妈妈用品']
    # 遍历每一行
    for i in sheet1.iter_rows():
        # 遍历每一行中的每一列（也即每一行中的每个单元格）
        for j in i:
            print(i[(j.column) - 1].value, end='\t')
        print()
    '''
类别	品牌1	尺寸/规格	数量	实付总价（元）	渠道
护理类	十月结晶 - 护理垫+刀纸	12片/包（60*90cm）+1kg	1	54.9	淘宝
None	十月结晶 - 计量卫生巾	2包/份 - 3片/包	2	89.8	淘宝
None	十月结晶 - 安睡裤	6包/份 - 2片/包	1	34	淘宝
None	十月结晶 - 护理垫	12片/包 - 60*90cm	1	28.8	京东
None	十月结晶 - 一次性纯棉内裤	5条/盒 - XXL码	4	58.8	京东
None	高洁丝 - 日用卫生巾	10片/包 - 240mm	6	55.62	京东
None	高洁丝 - 夜用卫生巾1	8片/包 - 280mm	2	None	None
None	高洁丝 - 夜用卫生巾2	4片/包 - 420mm	3	None	None
None	十月结晶 - 私处清洗器	1个/盒	1	29.9	淘宝
None	洁丽雅 - 一次性马桶坐垫	60片/盒	1	27.9	京东（原先买的）
None	维达 - 普通湿巾	5包/袋	2	29.8	京东
None	维达 - 酒精湿巾	5包/袋	1	13.2	京东
哺乳类用品	matyz - 硅胶吸奶器	1个/盒	1	15.84	京东
None	新贝 - 防溢乳垫	100片/包	1	19.9	京东
None	兰思诺 - 羊毛脂膏	40ml/管	1	91.8	京东
衣帽类	嫚熙 - 夏季纯棉两件套睡衣（带哺乳口）	L码	1	289.9	淘宝
None	嫚熙 - 四季款纯棉两件套睡衣（带哺乳口）	XL码	1	249.9	淘宝
None	嫚熙 - 哺乳内衣	1个/袋 - L码	2	249.8	淘宝
None	嫚熙 - 夏季纯棉月子帽	1个/袋 - 均码	1	55.91	淘宝
None	十月结晶 - 月子袜	4双/包 - 均码	1	37.9	淘宝
其他生活用品	克洛伊 - 吸管水杯	500ml	1	59.8	京东
None	None	None	None	=SUM(E2:E22)	None
None	None	None	None	None	None
None	None	None	None	None	None
None	None	None	None	None	None
    '''
def test5():
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')
    sheet = file1['妈妈用品']
    sum = 0
    for i in sheet.iter_rows():
        for cell in i:
            if i[(cell.column) - 1].value == '护理类':
                m = i[4].value
                sum = sum + m
    print(sum)  # 533.96
