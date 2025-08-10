#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''=================================================
@Author ：kw
@IDE    ：PyCharm
@Date   ：2025/2/7 15:57
=================================================='''
import openpyxl  # 导入对应的库 —— openpyxl


def test():
    file1 = openpyxl.load_workbook('../../resources/data1.xlsx')  # 打开指定的.xlsx文件
    sheet1 = file1['妈妈用品']  # 打开.xlsx文件中指定的表
    list1 = []
    '''
    openpyxl 库并不支持直接复制单元格对象从一个工作表到另一个工作表。当你尝试将一个工作表中的单元格对象直接添加到另一个工作表时，openpyxl 会抛出“cells cannot be copied from other worksheets”的错误。
    这是因为 openpyxl 的设计不允许跨工作表复制单元格对象，需要设置values_only=True参数用于仅返回单元格的值而不包括其他信息
    '''
    for row in sheet1.iter_rows(values_only=True):  # 按行遍历.xlsx文件（values_only=True参数用于仅返回单元格的值而不包括其他信息）
        list1.append(row)  # 将遍历到的每一行作为一个数据添加到list列表
    print(f'读取的文件内容是{list1}:类型是{type(list1)}')

    workbook2 = openpyxl.Workbook()  # 创建一个新的工作簿对象
    worksheet2 = workbook2.active  # 激活当前sheet工作簿对象
    for row in list1:
        worksheet2.append(row)  # 将list列表中的数据加入到sheet工作簿对象
    workbook2.save('../resources/abc/file1.xlsx')  # 保存工作簿到文件


